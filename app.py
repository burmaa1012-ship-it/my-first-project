from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
import re

app = Flask(__name__)
app.secret_key = 'register-secret-key-change-this'

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / 'Register.xlsx'
EMAIL_PATTERN = re.compile(r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$')
PHONE_PATTERN = re.compile(r'^[0-9+\-\s]{8,20}$')


def initialize_excel_file() -> None:
    if not EXCEL_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'Register'
        ws.append(['Огноо', 'Овог', 'Нэр', 'Утасны дугаар', 'Имэйл'])
        wb.save(EXCEL_FILE)


def save_registration(last_name: str, first_name: str, phone: str, email: str) -> None:
    initialize_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb['Register']
    ws.append([
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        last_name,
        first_name,
        phone,
        email,
    ])
    wb.save(EXCEL_FILE)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        last_name = request.form.get('last_name', '').strip()
        first_name = request.form.get('first_name', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()

        if not all([last_name, first_name, phone, email]):
            flash('Бүх талбарыг бөглөнө үү.', 'error')
            return render_template('index.html', form=request.form)

        if not EMAIL_PATTERN.match(email):
            flash('Имэйл хаяг буруу байна. Жишээ: name@example.com', 'error')
            return render_template('index.html', form=request.form)

        if not PHONE_PATTERN.match(phone):
            flash('Утасны дугаар буруу байна. Зөвхөн тоо, зай, +, - тэмдэгт ашиглана уу.', 'error')
            return render_template('index.html', form=request.form)

        save_registration(last_name, first_name, phone, email)
        flash('Бүртгэл амжилттай хийгдлээ.', 'success')
        return redirect(url_for('index'))

    return render_template('index.html', form={})


if __name__ == '__main__':
    initialize_excel_file()
    app.run(debug=True)
